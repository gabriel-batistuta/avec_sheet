# -*- coding: utf-8 -*-

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
import re
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from time import sleep
import time
import json
import os
import os.path
import platform
import modules as md
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import OperationSystemManager
from webdriver_manager.core.os_manager import ChromeType
from datetime import datetime, timedelta
import requests
from tests import sheet
import sys

sys.stdout.reconfigure(encoding='utf-8')

BOLD = "\033[1m"
BLACK  = "\033[0;30m"
RED    = "\033[0;31m"
GREEN  = "\033[0;32m"
YELLOW = "\033[0;33m"
BLUE   = "\033[0;34m"
MAGENTA = "\033[0;35m"
CYAN   = "\033[0;36m"
WHITE  = "\033[0;37m"

BOLD_BLACK  = "\033[1;30m"
BOLD_RED    = "\033[1;31m"
BOLD_GREEN  = "\033[1;32m"
BOLD_YELLOW = "\033[1;33m"
BOLD_BLUE   = "\033[1;34m"
BOLD_MAGENTA = "\033[1;35m"
BOLD_CYAN   = "\033[1;36m"
BOLD_WHITE  = "\033[1;37m"

RESET = "\033[0;0m"

def login_site():
    
    print('fazendo login...')
#     requests.post("http://localhost:3000", data="fazendo login...")
    login_url = 'https://admin.avec.beauty/somestore/admin'
    driver.get(login_url)
    # sleep(6)
    sleep(20)
    # email_field = WebDriverWait(driver, 20).until(
        # EC.visibility_of_element_located((By.XPATH, '//*[@id="formEmail"]'))
    # )
    email_field = driver.find_element(By.XPATH, '//*[@id="formEmail"]')
    email_field.send_keys(credentials['email'])
    password_field = driver.find_element(By.XPATH, '//*[@id="formSenha"]')
    password_field.send_keys(credentials['password'])
    button_submit = driver.find_element(By.CSS_SELECTOR, 'button.btn')
    button_submit.click()

def get_max_option_in_select():
    try:
        select_elem = driver.find_element(By.NAME, 'tableFilter_length')
        select = Select(select_elem)
        select.select_by_value("500")
    except:
        print('error in select 500 values')
#         requests.post("http://localhost:3000", data="error in select 500 values", )

def get_infos_of_general_report_page():
    def get_report_header():
        reports_header = []
        thead = driver.find_element(By.TAG_NAME, 'thead')
        report_blocks_head = thead.find_elements(By.TAG_NAME, 'th')
        report_blocks_head.pop() # retirando ultimo elemento de click pra página
        for i in report_blocks_head:
            reports_header.append(i.text.strip())
        return reports_header

    def get_values_of_header(reports_header):
        reports = []

        tbody = driver.find_element(By.TAG_NAME, 'tbody')
        blocks_reports = tbody.find_elements(By.TAG_NAME, 'tr')
        for block in blocks_reports:
            cells = block.find_elements(By.TAG_NAME, 'td')
            # Supondo que os primeiros N cells correspondam aos cabeçalhos em reports_header
            text_cells = cells[:-1]  # Todas as células exceto a última (que contém os links)
            link_cell = cells[-1]     # Célula com os links

            # Cria um dicionário associando cada cabeçalho ao valor de texto da célula correspondente
            report_data = {header: cell.text.strip() for header, cell in zip(reports_header, text_cells)}

            # Trata a célula de links: se houver mais de um <a>, escolhe o segundo, senão o primeiro
            anchors = link_cell.find_elements(By.TAG_NAME, 'a')
            if anchors:
                report_data['url'] = anchors[1].get_attribute('href') if len(anchors) > 1 else anchors[0].get_attribute('href')
            else:
                report_data['url'] = None

            reports.append(report_data)

            # temporário
            # break

        
        reports_dict = {
            "reports":reports
        }

        return reports_dict
    
    sleep(2)
    reports_header = get_report_header()
    reports_dict = get_values_of_header(reports_header)
    return reports_dict

def filter_lists(reports):
    
    print('filtrando categorias...')
#     requests.post("http://localhost:3000", data="filtrando categorias...", )
    categories = []
    lists = {}
    for r in reports:        
        print(f'R: ', r)
        # if r['Categoria'] not in categories:
            # print(f'categoria: {r["Categoria"]}')
        categories.append(r['Categoria'])
    for c in categories:
        lists[c] = []
        for report in reports:
            print(f'report: {c}')
            if report['Categoria'] == c:
                lists[c].append(report)
    make_json('reports_lists.json', lists)
    return lists

def table_to_dict(headers, values):
    data = {}
    for header, value in zip(headers, values):
        data[header] = value
    return data

def set_only_values_between(last_date_updated=None):
    MAX_ATTEMPTS = 4

    def get_start_date(months_ago: int):
        today = datetime.now()
        target_month = today.month - months_ago
        target_year = today.year

        while target_month <= 0:
            target_month += 12
            target_year -= 1

        start_date = datetime(target_year, target_month, 1)
        return start_date.strftime("%d/%m/%Y")

    if last_date_updated is None:
        last_date_updated = '01/01/2020'
    atual_date = datetime.now().strftime("%d/%m/%Y")

    sleep(3)
    try:
        field_initial_date = driver.find_element(By.NAME, 'inicio')
        field_final_date = driver.find_element(By.NAME, 'fim')
        search_button = driver.find_element(By.CSS_SELECTOR, '.btn-info')
    except Exception as e:
        print(f"Erro ao encontrar os campos: {e}. provavel página sem campo de data")
#         requests.post("http://localhost:3000", data=f"Erro ao encontrar os campos: {e}. provavel página sem campo de data")
        try:
            search_button = driver.find_element(By.CSS_SELECTOR, '.btn-info')
            search_button.click()
            print("Esperando carregar dados...")
#             requests.post("http://localhost:3000", data="Esperando carregar dados...")
            sleep(30)

            tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
            tags_rows = tbody.find_elements(By.TAG_NAME, 'td')

            # Se o único td encontrado indicar que o relatório está vazio
            if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
                print("Relatório sem dados, foi pulado.")
#                 requests.post("http://localhost:3000", data="Relatório sem dados, foi pulado.")
                return None
            else:
                return True
        except:
            return None

    print(f"Data inicial: {last_date_updated}")
#     requests.post("http://localhost:3000", data=f"Data inicial: {last_date_updated}")
    print(f"Data final: {atual_date}")
#     requests.post("http://localhost:3000", data=f"Data final: {atual_date}")

    for attempt in range(1, MAX_ATTEMPTS + 1):
        print(f"Tentativa {attempt} de {MAX_ATTEMPTS}")
#         requests.post("http://localhost:3000", data=f"Tentativa {attempt} de {MAX_ATTEMPTS}")

        # Insere os valores nos campos de data
        driver.execute_script("arguments[0].value = arguments[1];", field_initial_date, last_date_updated)
        driver.execute_script("arguments[0].value = arguments[1];", field_final_date, atual_date)

        initial_date_value = field_initial_date.get_attribute('value')
        final_date_value = field_final_date.get_attribute('value')

        print(f"Valor no campo inicial: {initial_date_value}")
#         requests.post("http://localhost:3000", data=f"Valor no campo inicial: {initial_date_value}")
        print(f"Valor no campo final: {final_date_value}")
#         requests.post("http://localhost:3000", data=f"Valor no campo final: {final_date_value}")

        if initial_date_value == last_date_updated and final_date_value == atual_date:
            try:
                search_button.click()
            except Exception as e:
                print(f"Erro ao clicar no botão de busca: {e}")
#                 requests.post("http://localhost:3000", data=f"Erro ao clicar no botão de busca: {e}")
                sleep(5)
                continue  # Tenta novamente

            sleep(30)
            print("Esperando carregar dados...")
#             requests.post("http://localhost:3000", data="Esperando carregar dados...")

            try:
                tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
                tags_rows = tbody.find_elements(By.TAG_NAME, 'td')

                # Se o único td encontrado indicar que o relatório está vazio
                if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
                    print("Relatório sem dados, foi pulado.")
#                     requests.post("http://localhost:3000", data="Relatório sem dados, foi pulado.")
                    return None
                else:
                    return True
            except Exception as e:
                print(f"Erro ao processar os dados: {e}")
#                 requests.post("http://localhost:3000", data=f"Erro ao processar os dados: {e}")
        else:
            print("Os valores dos campos não foram atualizados corretamente.")
#             requests.post("http://localhost:3000", data="Os valores dos campos não foram atualizados corretamente.")

        sleep(5)  # Aguarda alguns segundos antes de tentar novamente

    print("Número máximo de tentativas atingido. Falha ao atualizar os valores.")
#     requests.post("http://localhost:3000", data="Número máximo de tentativas atingido. Falha ao atualizar os valores.")
    return None

            
def get_infos_in_report_page(link_report_page, relatorie, last_date_updated=None):

    def get_headers():
        head_titles = []
        div_thead = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '#tableFilter > thead:nth-child(1) > tr:nth-child(1)'))
        )
        div_thead = driver.find_element(By.CSS_SELECTOR, '#tableFilter > thead:nth-child(1) > tr:nth-child(1)')
        tags_head = div_thead.find_elements(By.TAG_NAME, 'th')
        for tag in tags_head:
            title_head = tag.find_element(By.TAG_NAME, 'div').text
            head_titles.append(title_head)
        return head_titles
        
    def tables_to_dict(headers, values):
        data_list = []
        num_headers = len(headers)
        num_values = len(values)
        group_size = num_values // num_headers
        
        for i in range(0, num_values, num_headers):
            group_values = values[i:i+num_headers]
            data = {}
            for header, value in zip(headers, group_values):
                data[header] = value
            data_list.append(data)
        
        return data_list 

    def get_value_of_headers(headers, relatorie):
        values = []
        sleep(12)
        tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
        tags_rows = tbody.find_elements(By.TAG_NAME, 'td')
        if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
            print('relatorio sem dados, foi pulado')
#             requests.post("http://localhost:3000", data=f"relatorio sem dados, foi pulado", )
            return None
        while True:
            tbody = driver.find_element(By.CSS_SELECTOR, '#tableFilter > tbody:nth-child(2)')
            tags_rows = tbody.find_elements(By.TAG_NAME, 'td')
            # sleep(2)
            # if len(tags_rows) == 1 and tags_rows[0].get_attribute('class') == 'dataTables_empty':
            #     print('pulado')
            #     return None
            for tag in tags_rows:
                values.append(tag.text.strip())

            next_tab = driver.find_element(By.CSS_SELECTOR, 'li.next')
            # print(next_tab.text)
            try:
                class_name = next_tab.get_attribute('class')

                if 'disabled' in str(class_name):
                    # print('sem mais páginas pra pegar')
                    break
                else:
                    a_tag = next_tab.find_element(By.TAG_NAME, 'a')
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    #sleep(1)
                    # print('conseguiu pegar o link do proximo')
                    try:
                        a_tag.click()
                    except:
                        # print('erro ao clicar no proximo elemento')
                        break
            except NoSuchElementException:
                
                print('proximo elemento não encontrado')
                break
        # print(f'values: {len(values)}')
        print(f'número de valores encontrados: {GREEN}{int(len(values)/len(headers))}{RESET}')
#         requests.post("http://localhost:3000", data=f"número de valores encontrados: {int(len(values)/len(headers))}", )
        list_result = tables_to_dict(headers, values)
        dict_result = {
            relatorie:list_result
        }

        return dict_result
    
    # test com maior quantidade de dados possivel
    # link_report_page='https://admin.avec.beauty/admin/relatorio/0125'
    # login_site()
    driver.get(link_report_page)
    
    print(f'{YELLOW}{link_report_page}{RESET}')
#     requests.post("http://localhost:3000", data=f"{link_report_page}", )
    if last_date_updated is not None:
        response = set_only_values_between(last_date_updated)
    else:
        response = set_only_values_between()
    if response is None:
        return None
    get_max_option_in_select()
    headers = get_headers()
    report_dict = get_value_of_headers(headers, relatorie)

    return report_dict

def get_reports(last_date_updated=None):
    def get_lists_json():
        print('buscando e ordenando links...')
#         requests.post("http://localhost:3000", data=f"buscando e ordenando links...", )

        driver.get('https://admin.avec.beauty/admin/relatorios')
        print(f'{YELLOW}https://admin.avec.beauty/admin/relatorios{RESET}')
#         requests.post("http://localhost:3000", data=f"link dos relatórios: https://admin.avec.beauty/admin/relatorios", )
        sleep(5)
        get_max_option_in_select()
        reports_dict = get_infos_of_general_report_page()
        reports = reports_dict['reports']
        reports_lists = filter_lists(reports)
        return reports_lists
    reports_lists = get_lists_json()
    make_json('reports_lists.json',reports_lists)
    
    # temporario 
    # reports_lists = load_json('reports_lists.json')

    reports = []

    if last_date_updated is not None:
        print(f'Buscando relatórios de {last_date_updated} até {datetime.now().strftime("%d/%m/%Y")}')
#         requests.post("http://localhost:3000", data=f"Buscando relatórios de {last_date_updated} até {datetime.now().strftime('%d/%m/%Y')}", )

    def get_infos(reports_lists):
        # buscar apenas informações de relatórios de uma categoria especifica
        priority_category_list = ['Profissionais', 'Financeiro']

        def push_report(report):
            if report['Categoria'] not in priority_category_list:
                return None
            # if 'CUSTOS' not in report['Relatório'].upper() and report['Categoria'] != 'Financeiro':
                # return None
            report_dict = get_infos_in_report_page(report['url'], report['Relatório'], last_date_updated)
            if report_dict is not None:
                return report_dict
            else:
                return None
        
       
        for j, (category, reports_list) in enumerate(reports_lists.items()):
            category_reports = []
            priority_category_list = ['Financeiro', 'Profissionais']
            if category not in priority_category_list:
                continue
            for i, report in enumerate(reports_list):
                report_dict = push_report(report)
                if report_dict is not None:
                    category_reports.append(report_dict)
            if category_reports:
                reports_by_category = {
                    category: category_reports
                }
                reports.append(reports_by_category)

    get_infos(reports_lists)
    make_json('reports.json', {'reports':reports})

    # temporario
    # reports = load_json('reports.json')
    # return reports

    return {'reports':reports}

def make_json(file_name, dict):
    with open(file_name, 'w', encoding="utf8") as file:
        json.dump(dict, file, indent=4, ensure_ascii=False)

def load_json(file_name):
    with open(file_name, 'r', encoding="utf8") as file:
        return json.load(file)

def find_variable_name(value):
    for name, val in globals().items():
        if val is value:
            return name
    return None

def make_excel_table(reports:dict):
    folder_name = './planilhas'

    def make_folder():
        
        print('criando pasta de planilhas...')
#         requests.post("http://localhost:3000", data=f"criando pasta de planilhas...", )
        if os.path.isdir('./planilhas'):
            pass
        else:
            os.mkdir('./planilhas')
    
    def values_already_exists(book_name, sheet_name, values_to_insert):
        
        print('verificando existência de valores nas planilhas...')
#         requests.post("http://localhost:3000", data=f"verificando existência de valores nas planilhas...", )
        book = openpyxl.load_workbook(book_name)
        sheet = book[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            if list(row) == values_to_insert:
                    
                    print('valor já existe, não adicionado')
                    return True
            else:
                continue

    import openpyxl

    def update_value_in_excel(excel_file):
        # Abrir o arquivo Excel
        workbook = openpyxl.load_workbook(excel_file)
        
        # Iterar por todas as planilhas
        for sheet in workbook.worksheets:
            # Listar para armazenar as linhas modificadas
            modified_rows = []
            
            # Iterar por todas as linhas e células
            for row in sheet.iter_rows(values_only=True):
                new_row = []
                for cell in row:
                    if isinstance(cell, str):
                        if cell.count('"') == 1:
                            new_value = cell.replace('"', '')
                        elif cell.endswith('.0') and cell.count('.0') == 1:
                            new_value = int(cell.replace('.0', ''))
                        elif cell.isdigit():
                            new_value = int(cell)
                        else:
                            new_value = cell
                    else:
                        new_value = cell
                    new_row.append(new_value)
                modified_rows.append(new_row)
            
            # Atualizar as células da planilha com os novos valores
            for i, row in enumerate(modified_rows, start=1):
                for j, value in enumerate(row, start=1):
                    sheet.cell(row=i, column=j, value=value)
        
        # Salvar o arquivo Excel modificado
        workbook.save(excel_file)

    def make_excel(reports, folder_name):
        
        print('criando planilhas...')
        reports = reports['reports']
        for categories_items in reports:
            for category_name, values in categories_items.items():
                book = openpyxl.Workbook()
                for report in values:
                    for key, value in report.items():
                        title = re.sub(INVALID_TITLE_REGEX, '_', key)
                        if title != 'category':
                            book.create_sheet(title)
                        categorie_page = book[title]
                        headers_added = False
                        for item in value:
                            headers = []
                            tuples = []
                            if type(item) != dict:
                                
                                print(item)
                                continue
                            for key_, value_ in item.items():
                                headers.append(key_)
                                tuples.append(value_) 
                            if headers_added is False:
                                categorie_page.append(headers)
                                headers_added = True
                            categorie_page.append(tuples)
                pattern_sheet = book['Sheet']
                book.remove(pattern_sheet)
                book.save(f'{folder_name}/{category_name}.xlsx')
                update_value_in_excel(f'{folder_name}/{category_name}.xlsx')
                
                print(f'planilha: {category_name}.xlsx criada')

    make_folder()
    make_excel(reports, folder_name)
    md.send_sheets()
    md.change_values()
    
if __name__ == '__main__':
    # headers do post pra ir texto utf-8
    start_time = time.time()
    with open('settings.json', 'r', encoding='utf-8') as file:
        credentials = json.load(file)["credentials"]

    chrome_options = Options()
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('log-level=3')
    if platform.system() == 'Linux':
        # chrome_driver_path = Service(ChromeDriverManager().install())
        try:
            chrome_driver_path = print(ChromeDriverManager(os_system_manager=OperationSystemManager(os_type='linux64')).install())
            driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
        except:
            chrome_driver_path = Service('drivers/chromedriver')
            driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
    else:
        try:
            chrome_driver_path = print(ChromeDriverManager(os_system_manager=OperationSystemManager(os_type='win64')).install())
            driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
        except:
            chrome_driver_path = Service('drivers/chromedriver_131_WINDOWS.exe')
            driver = webdriver.Chrome(service=chrome_driver_path, options=chrome_options)
    wait = WebDriverWait(driver, 5)
    if os.path.isfile('./settings.json'):
        settings = load_json('./settings.json')
        login_site()
        reports = get_reports(settings['last_date_updated'])
    else:
        login_site()
        reports = get_reports()
    
    make_excel_table(reports)

    # final do programa
    JSON = load_json('settings.json')
    
    print(JSON)
    JSON['last_date_updated'] = datetime.now().strftime("%d/%m/%Y")
    make_json('settings.json', JSON)
    end_time = time.time()
    duration = (end_time - start_time) / 60
    
    print("Tempo decorrido: ", duration, "minutos")
#     requests.post("http://localhost:3000", data=f"Tempo decorrido: {duration} minutos", )
    # with open('reports.json', 'r') as file:
        # reports = json.load(file)
    # make_excel_table(reports)